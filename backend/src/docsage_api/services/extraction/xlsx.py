"""XLSX extraction: one markdown table per non-empty sheet (used range only)."""

from pathlib import Path

from openpyxl import load_workbook

from docsage_api.services.extraction.base import ExtractedPart, ExtractionResult, markdown_table

MAX_ROWS_PER_SHEET = 200


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))  # 12.0 -> "12"
    return str(value)


def extract(path: Path) -> ExtractionResult:
    workbook = load_workbook(filename=str(path), data_only=True)
    result = ExtractionResult()

    for sheet in workbook.worksheets:
        rows = [[_cell(c) for c in row] for row in sheet.iter_rows(values_only=True)]
        while rows and not any(cell.strip() for cell in rows[-1]):
            rows.pop()  # trim trailing empty rows of the used range
        if not any(any(cell.strip() for cell in row) for row in rows):
            continue  # skip empty sheets

        truncated = len(rows) > MAX_ROWS_PER_SHEET
        if truncated:
            rows = rows[:MAX_ROWS_PER_SHEET] + [["…truncated"]]
        result.parts.append(
            ExtractedPart(
                kind="table",
                content=f"### Sheet: {sheet.title}\n\n{markdown_table(rows)}",
            )
        )
    workbook.close()
    return result
